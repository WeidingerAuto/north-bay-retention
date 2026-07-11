"""
Import all data from the Excel lease retention spreadsheet into PostgreSQL.
Safe to re-run — skips rows already in the database.

Usage:
    python scripts/migrate_excel.py
"""

import os
import re
import psycopg2
import openpyxl

EXCEL_PATH = r"\\nb-server1\SalesManagment\Lease Retention\temp lease term info.xlsx"

DB_URL = os.environ.get("DATABASE_URL")
if not DB_URL:
    raise SystemExit("Set DATABASE_URL before running (see cred-rotation runbook).")

MONTH_MAP = {
    "jan": "Jan", "feb": "Feb", "mar": "Mar", "apr": "Apr",
    "may": "May", "jun": "Jun", "june": "Jun",
    "jul": "Jul", "july": "Jul",
    "aug": "Aug", "august": "Aug",
    "sep": "Sep", "sept": "Sep",
    "oct": "Oct", "nov": "Nov", "dec": "Dec",
    "april": "Apr",
}


def parse_sheet_name(name):
    """Return (month_abbr, year) or None if not a data sheet."""
    m = re.match(r"^(\w+)\s+(\d{4})$", name.strip())
    if not m:
        return None
    mon = MONTH_MAP.get(m.group(1).lower())
    if not mon:
        return None
    return mon, int(m.group(2))


def cell_val(row, idx):
    v = row[idx] if idx < len(row) else None
    return str(v).strip() if v is not None else ""


def detect_outcome(row):
    """Return A, B, or C based on which column has an X."""
    for col, outcome in [(1, "A"), (2, "B"), (3, "C")]:
        v = cell_val(row, col).upper()
        if v and v not in ("NONE", ""):
            return outcome
    return None


def main():
    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    inserted = skipped_dupe = skipped_invalid = 0

    for sheet_name in wb.sheetnames:
        parsed = parse_sheet_name(sheet_name)
        if not parsed:
            continue
        month, year = parsed
        ws = wb[sheet_name]

        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            name = cell_val(row, 0).strip()
            if not name:
                continue

            outcome = detect_outcome(row)
            if not outcome:
                skipped_invalid += 1
                continue

            disposition = cell_val(row, 4).strip()
            broker = cell_val(row, 6).strip()

            # Idempotent: skip if identical row already exists
            cur.execute(
                """SELECT 1 FROM lease_entries
                   WHERE name=%s AND outcome=%s AND year=%s AND month=%s""",
                (name, outcome, year, month),
            )
            if cur.fetchone():
                skipped_dupe += 1
                continue

            cur.execute(
                """INSERT INTO lease_entries
                       (name, outcome, disposition, broker, our_customer, year, month, is_historical)
                   VALUES (%s, %s, %s, %s, TRUE, %s, %s, %s)""",
                (name, outcome, disposition, broker, year, month, year < 2026),
            )
            inserted += 1

    conn.commit()
    cur.close()
    conn.close()
    wb.close()

    print(f"Done.")
    print(f"  Inserted : {inserted}")
    print(f"  Dupes    : {skipped_dupe}")
    print(f"  No A/B/C : {skipped_invalid}")


if __name__ == "__main__":
    main()
